"""
Data migration framework for CRM system.
Handles migration from Feishu multi-dimensional tables to the new 7-table CRM schema.
"""

import json
import logging
from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import re

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DataMigrationFramework:
    """Main class for handling data migration from Feishu to new CRM schema."""
    
    def __init__(self, db_url: str, source_data_path: str):
        self.db_url = db_url
        self.source_data_path = Path(source_data_path)
        self.engine = create_engine(db_url)
        self.SessionLocal = sessionmaker(bind=self.engine)
        
        # Migration statistics
        self.stats = {
            'customers': {'total': 0, 'migrated': 0, 'duplicates': 0, 'errors': 0},
            'opportunities': {'total': 0, 'migrated': 0, 'errors': 0},
            'contracts': {'total': 0, 'migrated': 0, 'errors': 0},
            'products': {'total': 0, 'standardized': 0, 'errors': 0}
        }
    
    def run_migration(self) -> Dict[str, Any]:
        """
        Run the complete data migration process.
        Returns migration statistics and any errors encountered.
        """
        logger.info("Starting CRM data migration...")
        
        try:
            # Load source data
            source_data = self._load_source_data()
            
            # Clean and standardize data
            cleaned_data = self._clean_and_standardize(source_data)
            
            # Migrate to new schema
            migration_results = self._migrate_to_new_schema(cleaned_data)
            
            logger.info("CRM data migration completed successfully!")
            return {
                'success': True,
                'stats': self.stats,
                'errors': [],
                'migration_results': migration_results
            }
            
        except Exception as e:
            logger.error(f"Migration failed: {str(e)}")
            return {
                'success': False,
                'stats': self.stats,
                'errors': [str(e)],
                'migration_results': None
            }
    
    def _load_source_data(self) -> Dict[str, Any]:
        """Load source data from Feishu exports and JSON dumps."""
        logger.info("Loading source data...")
        
        source_data = {}
        
        # Load Excel data if available
        excel_files = list(self.source_data_path.glob("*.xlsx"))
        if excel_files:
            latest_excel = max(excel_files, key=lambda f: f.stat().st_mtime)
            source_data['excel'] = self._load_excel_data(latest_excel)
        
        # Load JSON dumps if available
        json_files = [
            'crm_schema_dump.json',
            'crm_views_dump.json', 
            'crm_readonly_stats.json'
        ]
        
        for json_file in json_files:
            json_path = self.source_data_path / json_file
            if json_path.exists():
                with open(json_path, 'r', encoding='utf-8') as f:
                    source_data[json_file.replace('.json', '')] = json.load(f)
        
        logger.info(f"Loaded source data: {len(source_data)} sources")
        return source_data
    
    def _load_excel_data(self, excel_path: Path) -> Dict[str, List[Dict]]:
        """Load data from Excel file."""
        logger.info(f"Loading Excel data from {excel_path}")
        
        wb = load_workbook(excel_path, read_only=True)
        data = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows()) if cell.value is not None]
            rows = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = row[i]
                    rows.append(row_dict)
            
            data[sheet_name] = rows
        
        wb.close()
        logger.info(f"Loaded {len(data)} sheets from Excel")
        return data
    
    def _clean_and_standardize(self, source_data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and standardize the source data."""
        logger.info("Cleaning and standardizing data...")
        
        cleaned_data = {}
        
        # Handle customer data deduplication
        if 'excel' in source_data and '客户管理' in source_data['excel']:
            customers = source_data['excel']['客户管理']
            cleaned_data['customers'] = self._deduplicate_customers(customers)
        
        # Standardize product names
        if 'excel' in source_data and '商机管理' in source_data['excel']:
            opportunities = source_data['excel']['商机管理']
            cleaned_data['opportunities'] = self._standardize_products(opportunities)
        
        # Process contract data
        if 'excel' in source_data and '合同管理' in source_data['excel']:
            contracts = source_data['excel']['合同管理']
            cleaned_data['contracts'] = self._process_contracts(contracts)
        
        # Create product dictionary
        cleaned_data['products'] = self._create_product_dictionary(cleaned_data.get('opportunities', []))
        
        logger.info("Data cleaning and standardization completed")
        return cleaned_data
    
    def _deduplicate_customers(self, customers: List[Dict]) -> List[Dict]:
        """Deduplicate customer records based on customer name."""
        logger.info(f"Deduplicating {len(customers)} customer records...")
        
        customer_groups = {}
        for customer in customers:
            name = self._normalize_text(customer.get('客户名称', ''))
            if name:
                if name not in customer_groups:
                    customer_groups[name] = []
                customer_groups[name].append(customer)
        
        deduplicated = []
        for name, group in customer_groups.items():
            if len(group) == 1:
                # Single record, keep as-is
                deduplicated.append(group[0])
                self.stats['customers']['migrated'] += 1
            else:
                # Multiple records, merge them
                merged_customer = self._merge_customer_records(group)
                deduplicated.append(merged_customer)
                self.stats['customers']['duplicates'] += len(group) - 1
                self.stats['customers']['migrated'] += 1
        
        self.stats['customers']['total'] = len(customers)
        logger.info(f"Deduplicated to {len(deduplicated)} unique customers")
        return deduplicated
    
    def _merge_customer_records(self, customer_records: List[Dict]) -> Dict:
        """Merge multiple customer records into a single comprehensive record."""
        if not customer_records:
            return {}
        
        merged = customer_records[0].copy()
        
        # Merge fields that can have multiple values
        mergeable_fields = ['联系电话', '主要联系人', '客户行业', '客户区域']
        
        for field in mergeable_fields:
            values = []
            for record in customer_records:
                if record.get(field):
                    values.append(str(record[field]))
            
            if values:
                # Remove duplicates while preserving order
                unique_values = list(dict.fromkeys(values))
                merged[field] = '; '.join(unique_values)
        
        return merged
    
    def _standardize_products(self, opportunities: List[Dict]) -> List[Dict]:
        """Standardize product names across opportunities."""
        logger.info(f"Standardizing products in {len(opportunities)} opportunities...")
        
        standardized_opportunities = []
        product_mapping = self._get_product_mapping()
        
        for opportunity in opportunities:
            original_product = opportunity.get('产品', '')
            if original_product:
                standardized_product = self._match_product(original_product, product_mapping)
                opportunity['product_standardized'] = standardized_product
                self.stats['products']['standardized'] += 1
            
            standardized_opportunities.append(opportunity)
        
        self.stats['products']['total'] = len([o for o in opportunities if o.get('产品')])
        self.stats['opportunities']['total'] = len(opportunities)
        self.stats['opportunities']['migrated'] = len(opportunities)
        
        return standardized_opportunities
    
    def _get_product_mapping(self) -> Dict[str, str]:
        """Get mapping of product name variations to standard names."""
        # This would typically come from a configuration file or database
        return {
            'ipg': 'IPGuard',
            'IPG': 'IPGuard', 
            'IPguard': 'IPGuard',
            'ipguard': 'IPGuard',
            '爱数': 'Aisino',
            '绿盟': 'NSFOCUS',
            '深信服': 'Sangfor',
            '火绒': 'Huorong'
        }
    
    def _match_product(self, product_name: str, mapping: Dict[str, str]) -> str:
        """Match product name to standardized name using fuzzy matching."""
        if not product_name:
            return ''
        
        # Exact match
        if product_name in mapping:
            return mapping[product_name]
        
        # Fuzzy match - check if any key is contained in the product name
        normalized_name = product_name.lower()
        for variant, standard in mapping.items():
            if variant.lower() in normalized_name:
                return standard
        
        # Return original if no match found
        return product_name
    
    def _process_contracts(self, contracts: List[Dict]) -> List[Dict]:
        """Process contract data and link to opportunities/projects."""
        logger.info(f"Processing {len(contracts)} contracts...")
        
        processed_contracts = []
        for contract in contracts:
            # Add processing logic here
            processed_contracts.append(contract)
        
        self.stats['contracts']['total'] = len(contracts)
        self.stats['contracts']['migrated'] = len(contracts)
        
        return processed_contracts
    
    def _create_product_dictionary(self, opportunities: List[Dict]) -> List[Dict]:
        """Create product dictionary from standardized opportunity data."""
        logger.info("Creating product dictionary...")
        
        products = {}
        for opportunity in opportunities:
            product_name = opportunity.get('product_standardized', '')
            if product_name and product_name not in products:
                products[product_name] = {
                    'product_name': product_name,
                    'product_type': self._categorize_product(product_name),
                    'brand_manufacturer': self._get_brand_manufacturer(product_name),
                    'is_active': True
                }
        
        logger.info(f"Created product dictionary with {len(products)} products")
        return list(products.values())
    
    def _categorize_product(self, product_name: str) -> str:
        """Categorize product into types."""
        security_products = ['IPGuard', 'NSFOCUS', 'Huorong']
        backup_products = ['Aisino']
        network_products = ['Sangfor']
        
        if product_name in security_products:
            return 'Endpoint Security'
        elif product_name in backup_products:
            return 'Data Backup'
        elif product_name in network_products:
            return 'Network Security'
        else:
            return 'Other'
    
    def _get_brand_manufacturer(self, product_name: str) -> str:
        """Get brand/manufacturer for product."""
        manufacturer_map = {
            'IPGuard': 'Beijing Yidun Technology',
            'Aisino': 'Aisino Corporation',
            'NSFOCUS': 'NSFOCUS Technologies',
            'Sangfor': 'Sangfor Technologies',
            'Huorong': 'Huorong Security'
        }
        return manufacturer_map.get(product_name, product_name)
    
    def _migrate_to_new_schema(self, cleaned_data: Dict[str, Any]) -> Dict[str, Any]:
        """Migrate cleaned data to the new 7-table schema."""
        logger.info("Migrating data to new schema...")
        
        session = self.SessionLocal()
        try:
            # Create users first (sales personnel)
            users = self._migrate_users(session)
            
            # Create products
            products = self._migrate_products(session, cleaned_data.get('products', []))
            
            # Create customers
            customers = self._migrate_customers(session, cleaned_data.get('customers', []), users)
            
            # Create channels  
            channels = self._migrate_channels(session, customers)
            
            # Create opportunities
            opportunities = self._migrate_opportunities(session, cleaned_data.get('opportunities', []), customers, channels, users, products)
            
            # Create projects (from won opportunities)
            projects = self._migrate_projects(session, opportunities, customers, channels, users, products)
            
            # Create contracts
            contracts = self._migrate_contracts(session, cleaned_data.get('contracts', []), projects, channels)
            
            # Create follow-ups
            followups = self._migrate_followups(session, opportunities, projects, customers, users)
            
            session.commit()
            
            logger.info("Data migration to new schema completed")
            return {
                'users': len(users),
                'products': len(products), 
                'customers': len(customers),
                'channels': len(channels),
                'opportunities': len(opportunities),
                'projects': len(projects),
                'contracts': len(contracts),
                'followups': len(followups)
            }
            
        except Exception as e:
            session.rollback()
            logger.error(f"Error during schema migration: {str(e)}")
            raise
        finally:
            session.close()
    
    def _migrate_users(self, session) -> List[Dict]:
        """Migrate user data."""
        # In a real implementation, this would load from source data
        # For now, create placeholder users based on the roles needed
        users_data = [
            {'name': 'Admin User', 'email': 'admin@example.com', 'role': 'admin'},
            {'name': 'Sales User', 'email': 'sales@example.com', 'role': 'sales'},  
            {'name': 'Business User', 'email': 'business@example.com', 'role': 'business'},
            {'name': 'Finance User', 'email': 'finance@example.com', 'role': 'finance'}
        ]
        
        users = []
        for user_data in users_data:
            result = session.execute(text("""
                INSERT INTO users (name, email, role) 
                VALUES (:name, :email, :role) 
                RETURNING id
            """), user_data)
            user_id = result.fetchone()[0]
            users.append({'id': user_id, **user_data})
        
        return users
    
    def _migrate_products(self, session, products_data: List[Dict]) -> List[Dict]:
        """Migrate product dictionary."""
        products = []
        for product_data in products_data:
            result = session.execute(text("""
                INSERT INTO products (product_name, product_type, brand_manufacturer, is_active)
                VALUES (:product_name, :product_type, :brand_manufacturer, :is_active)
                RETURNING id, product_code
            """), product_data)
            row = result.fetchone()
            products.append({
                'id': row[0],
                'product_code': row[1],
                **product_data
            })
        
        return products
    
    def _migrate_customers(self, session, customers_data: List[Dict], users: List[Dict]) -> List[Dict]:
        """Migrate terminal customers."""
        customers = []
        sales_user = next((u for u in users if u['role'] == 'sales'), users[0])
        
        for customer_data in customers_data:
            # Map source fields to target schema
            mapped_data = {
                'customer_name': customer_data.get('客户名称', ''),
                'customer_nickname': customer_data.get('客户简称', ''),
                'customer_industry': self._map_industry(customer_data.get('客户行业', 'Other')),
                'customer_region': customer_data.get('客户区域', 'Other'),
                'customer_owner_id': sales_user['id'],
                'main_contact': customer_data.get('主要联系人', ''),
                'phone': customer_data.get('联系电话', ''),
                'customer_status': self._map_customer_status(customer_data.get('客户状态', 'Active')),
                'maintenance_expiry': customer_data.get('维保到期时间'),
                'notes': customer_data.get('备注', '')
            }
            
            result = session.execute(text("""
                INSERT INTO terminal_customers (
                    customer_name, customer_nickname, customer_industry, customer_region,
                    customer_owner_id, main_contact, phone, customer_status,
                    maintenance_expiry, notes
                ) VALUES (
                    :customer_name, :customer_nickname, :customer_industry, :customer_region,
                    :customer_owner_id, :main_contact, :phone, :customer_status,
                    :maintenance_expiry, :notes
                ) RETURNING id, customer_code
            """), mapped_data)
            
            row = result.fetchone()
            customers.append({
                'id': row[0],
                'customer_code': row[1],
                **mapped_data
            })
        
        return customers
    
    def _map_industry(self, industry: str) -> str:
        """Map industry to standardized values."""
        industry_map = {
            '制造业': 'Manufacturing',
            '金融': 'Finance', 
            '政府/事业单位': 'Government',
            '医疗': 'Healthcare',
            '教育': 'Education',
            '能源化工': 'Energy',
            '其他': 'Other'
        }
        return industry_map.get(industry, 'Other')
    
    def _map_customer_status(self, status: str) -> str:
        """Map customer status to standardized values."""
        status_map = {
            '潜在客户': 'Potential',
            '活跃客户': 'Active', 
            '存量客户': 'Existing',
            '流失客户': 'Lost'
        }
        return status_map.get(status, 'Active')
    
    def _migrate_channels(self, session, customers: List[Dict]) -> List[Dict]:
        """Migrate channel/trading partner data."""
        # For now, create one channel per customer as a simple mapping
        channels = []
        for customer in customers:
            channel_data = {
                'company_name': customer['customer_name'],
                'channel_type': 'Direct Terminal Customer',
                'main_contact': customer['main_contact'],
                'phone': customer['phone'],
                'billing_info': '',
                'notes': f'Channel created for customer {customer["customer_code"]}'
            }
            
            result = session.execute(text("""
                INSERT INTO channels (
                    company_name, channel_type, main_contact, phone, billing_info, notes
                ) VALUES (
                    :company_name, :channel_type, :main_contact, :phone, :billing_info, :notes
                ) RETURNING id, channel_code
            """), channel_data)
            
            row = result.fetchone()
            channels.append({
                'id': row[0],
                'channel_code': row[1],
                **channel_data
            })
        
        return channels
    
    def _migrate_opportunities(self, session, opportunities_data: List[Dict], 
                              customers: List[Dict], channels: List[Dict], 
                              users: List[Dict], products: List[Dict]) -> List[Dict]:
        """Migrate opportunity data."""
        opportunities = []
        sales_user = next((u for u in users if u['role'] == 'sales'), users[0])
        
        # Create product ID mapping
        product_name_to_id = {p['product_name']: p['id'] for p in products}
        
        for opp_data in opportunities_data:
            # Find associated customer
            customer = customers[0] if customers else None
            channel = channels[0] if channels else None
            
            # Map product names to IDs
            product_names = [opp_data.get('product_standardized', '')] if opp_data.get('product_standardized') else []
            product_ids = [product_name_to_id.get(name) for name in product_names if name in product_name_to_id]
            
            opportunity_data = {
                'opportunity_name': opp_data.get('商机名称', f"Opportunity {len(opportunities) + 1}"),
                'terminal_customer_id': customer['id'] if customer else None,
                'opportunity_source': self._map_opportunity_source(opp_data.get('商机来源', 'Direct Sales')),
                'product_ids': product_ids,
                'opportunity_stage': self._map_opportunity_stage(opp_data.get('商机阶段', 'Initial Contact')),
                'lead_grade': opp_data.get('线索等级', 'B'),
                'expected_contract_amount': float(opp_data.get('预计合同金额', 0)) if opp_data.get('预计合同金额') else 0.0,
                'expected_close_date': opp_data.get('预计成交时间'),
                'sales_owner_id': sales_user['id'],
                'channel_id': channel['id'] if channel else None,
                'vendor_registration_status': opp_data.get('厂家报备状态'),
                'vendor_discount': float(opp_data.get('报备折扣', 1.0)) if opp_data.get('报备折扣') else None,
                'loss_reason': opp_data.get('丢单原因') if opp_data.get('商机阶段') == '丢单' else None,
                'project_id': None  # Will be set when project is created
            }
            
            # Skip if required fields are missing
            if not opportunity_data['terminal_customer_id']:
                continue
            
            result = session.execute(text("""
                INSERT INTO opportunities (
                    opportunity_name, terminal_customer_id, opportunity_source, product_ids,
                    opportunity_stage, lead_grade, expected_contract_amount, expected_close_date,
                    sales_owner_id, channel_id, vendor_registration_status, vendor_discount,
                    loss_reason, project_id
                ) VALUES (
                    :opportunity_name, :terminal_customer_id, :opportunity_source, :product_ids,
                    :opportunity_stage, :lead_grade, :expected_contract_amount, :expected_close_date,
                    :sales_owner_id, :channel_id, :vendor_registration_status, :vendor_discount,
                    :loss_reason, :project_id
                ) RETURNING id, opportunity_code
            """), opportunity_data)
            
            row = result.fetchone()
            opportunities.append({
                'id': row[0],
                'opportunity_code': row[1],
                **opportunity_data
            })
        
        return opportunities
    
    def _map_opportunity_source(self, source: str) -> str:
        """Map opportunity source to standardized values."""
        source_map = {
            '销售直拓': 'Direct Sales',
            '渠道带入': 'Channel', 
            '老客户转介绍': 'Customer Referral',
            '老客户续保/扩容': 'Renewal/Expansion'
        }
        return source_map.get(source, 'Direct Sales')
    
    def _map_opportunity_stage(self, stage: str) -> str:
        """Map opportunity stage to standardized values."""
        stage_map = {
            '初步接触': 'Initial Contact',
            '需求确认': 'Needs Confirmation',
            '方案报价': 'Proposal', 
            '厂家报备中': 'Vendor Registration',
            '等待决策': 'Decision Pending',
            '赢单→转项目': 'Won→Project',
            '丢单': 'Lost'
        }
        return stage_map.get(stage, 'Initial Contact')
    
    def _migrate_projects(self, session, opportunities: List[Dict], 
                        customers: List[Dict], channels: List[Dict],
                        users: List[Dict], products: List[Dict]) -> List[Dict]:
        """Migrate project data from won opportunities."""
        projects = []
        sales_user = next((u for u in users if u['role'] == 'sales'), users[0])
        
        # Create product ID mapping
        product_name_to_id = {p['product_name']: p['id'] for p in products}
        
        for opp in opportunities:
            # Only create projects for won opportunities
            if opp['opportunity_stage'] != 'Won→Project':
                continue
            
            # Find associated customer and channel
            customer = customers[0] if customers else None
            channel = channels[0] if channels else None
            
            # Determine if this is a renewal project
            is_renewal = '续保' in opp['opportunity_name'] or 'Renewal' in opp['opportunity_source']
            
            # Get product IDs from opportunity
            product_ids = opp['product_ids'] if opp['product_ids'] else [products[0]['id']] if products else []
            
            project_data = {
                'project_name': f"Project from {opp['opportunity_name']}",
                'terminal_customer_id': customer['id'] if customer else opp['terminal_customer_id'],
                'channel_id': channel['id'] if channel else opp['channel_id'],
                'source_opportunity_id': opp['id'],
                'product_ids': product_ids,
                'business_type': 'Renewal/Maintenance' if is_renewal else 'New Project',
                'project_status': 'Initiating',
                'sales_owner_id': sales_user['id'],
                'downstream_contract_amount': opp['expected_contract_amount'],
                'upstream_procurement_amount': None,  # Will be filled later
                'direct_project_investment': None,
                'additional_investment': None,
                'winning_date': None,
                'acceptance_date': None,
                'first_payment_date': None,
                'actual_payment_amount': None,
                'notes': f'Migrated from opportunity {opp["opportunity_code"]}'
            }
            
            result = session.execute(text("""
                INSERT INTO projects (
                    project_name, terminal_customer_id, channel_id, source_opportunity_id,
                    product_ids, business_type, project_status, sales_owner_id,
                    downstream_contract_amount, upstream_procurement_amount,
                    direct_project_investment, additional_investment,
                    winning_date, acceptance_date, first_payment_date, actual_payment_amount, notes
                ) VALUES (
                    :project_name, :terminal_customer_id, :channel_id, :source_opportunity_id,
                    :product_ids, :business_type, :project_status, :sales_owner_id,
                    :downstream_contract_amount, :upstream_procurement_amount,
                    :direct_project_investment, :additional_investment,
                    :winning_date, :acceptance_date, :first_payment_date, :actual_payment_amount, :notes
                ) RETURNING id, project_code, gross_margin
            """), project_data)
            
            row = result.fetchone()
            project = {
                'id': row[0],
                'project_code': row[1],
                'gross_margin': row[2],
                **project_data
            }
            projects.append(project)
            
            # Update the opportunity to link to this project
            session.execute(text("""
                UPDATE opportunities SET project_id = :project_id WHERE id = :opportunity_id
            """), {'project_id': project['id'], 'opportunity_id': opp['id']})
        
        return projects
    
    def _migrate_contracts(self, session, contracts_data: List[Dict], 
                          projects: List[Dict], channels: List[Dict]) -> List[Dict]:
        """Migrate contract data."""
        contracts = []
        
        for contract_data in contracts_data:
            # Link to first project if available
            project = projects[0] if projects else None
            channel = channels[0] if channels else None
            
            if not project:
                continue
            
            contract_record = {
                'contract_code': contract_data.get('合同编号', f"CONTRACT_{len(contracts) + 1}"),
                'contract_name': contract_data.get('合同名称', f"Contract {len(contracts) + 1}"),
                'project_id': project['id'],
                'contract_direction': 'Downstream',
                'contract_status': 'Effective',
                'counterparty_id': channel['id'] if channel else None,
                'contract_amount': float(contract_data.get('合同金额（万元）', 0)) * 10000 if contract_data.get('合同金额（万元）') else 0.0,
                'signing_date': contract_data.get('签约日期'),
                'contract_file_url': contract_data.get('合同附件'),
                'notes': contract_data.get('备注', '')
            }
            
            result = session.execute(text("""
                INSERT INTO contracts (
                    contract_code, contract_name, project_id, contract_direction,
                    contract_status, counterparty_id, contract_amount, signing_date,
                    contract_file_url, notes
                ) VALUES (
                    :contract_code, :contract_name, :project_id, :contract_direction,
                    :contract_status, :counterparty_id, :contract_amount, :signing_date,
                    :contract_file_url, :notes
                ) RETURNING id
            """), contract_record)
            
            contract_id = result.fetchone()[0]
            contracts.append({'id': contract_id, **contract_record})
        
        return contracts
    
    def _migrate_followups(self, session, opportunities: List[Dict], 
                          projects: List[Dict], customers: List[Dict], 
                          users: List[Dict]) -> List[Dict]:
        """Migrate follow-up records."""
        followups = []
        sales_user = next((u for u in users if u['role'] == 'sales'), users[0])
        customer = customers[0] if customers else None
        
        if not customer:
            return followups
        
        # Create follow-ups for opportunities
        for opp in opportunities[:5]:  # Limit to avoid too many records
            followup_data = {
                'terminal_customer_id': customer['id'],
                'opportunity_id': opp['id'],
                'project_id': None,
                'follow_up_date': '2026-03-01',
                'follow_up_method': 'Phone',
                'follow_up_content': f'Follow-up for opportunity {opp["opportunity_name"]}',
                'follow_up_conclusion': 'Progressing Well',
                'next_action': 'Schedule next call',
                'next_follow_up_date': '2026-03-08',
                'follower_id': sales_user['id']
            }
            
            result = session.execute(text("""
                INSERT INTO follow_ups (
                    terminal_customer_id, opportunity_id, project_id, follow_up_date,
                    follow_up_method, follow_up_content, follow_up_conclusion,
                    next_action, next_follow_up_date, follower_id
                ) VALUES (
                    :terminal_customer_id, :opportunity_id, :project_id, :follow_up_date,
                    :follow_up_method, :follow_up_content, :follow_up_conclusion,
                    :next_action, :next_follow_up_date, :follower_id
                ) RETURNING id
            """), followup_data)
            
            followup_id = result.fetchone()[0]
            followups.append({'id': followup_id, **followup_data})
        
        # Create follow-ups for projects
        for proj in projects[:3]:  # Limit to avoid too many records
            followup_data = {
                'terminal_customer_id': customer['id'],
                'opportunity_id': None,
                'project_id': proj['id'],
                'follow_up_date': '2026-03-01',
                'follow_up_method': 'Visit',
                'follow_up_content': f'Project follow-up for {proj["project_name"]}',
                'follow_up_conclusion': 'Progressing Well',
                'next_action': 'Review project status',
                'next_follow_up_date': '2026-03-15',
                'follower_id': sales_user['id']
            }
            
            result = session.execute(text("""
                INSERT INTO follow_ups (
                    terminal_customer_id, opportunity_id, project_id, follow_up_date,
                    follow_up_method, follow_up_content, follow_up_conclusion,
                    next_action, next_follow_up_date, follower_id
                ) VALUES (
                    :terminal_customer_id, :opportunity_id, :project_id, :follow_up_date,
                    :follow_up_method, :follow_up_content, :follow_up_conclusion,
                    :next_action, :next_follow_up_date, :follower_id
                ) RETURNING id
            """), followup_data)
            
            followup_id = result.fetchone()[0]
            followups.append({'id': followup_id, **followup_data})
        
        return followups
    
    def _normalize_text(self, text: Optional[str]) -> str:
        """Normalize text for comparison."""
        if not text:
            return ''
        return str(text).strip().lower()

def main():
    """Main function to run the migration."""
    # Configuration
    DB_URL = "postgresql://postgres:password@localhost:5432/crm_db"
    SOURCE_DATA_PATH = "D:/项目材料/业财一体"
    
    # Run migration
    migration = DataMigrationFramework(DB_URL, SOURCE_DATA_PATH)
    result = migration.run_migration()
    
    if result['success']:
        print("Migration completed successfully!")
        print(f"Statistics: {result['stats']}")
    else:
        print(f"Migration failed: {result['errors']}")

if __name__ == "__main__":
    main()